""" Python Script to delete complete row having a strikethrough value"""
#Using openpyxl library which is used to read/write/delete values in excel.
#Install the library before using. (pip install openpyxl)
from openpyxl import load_workbook
#Change 'Test.xlsx' with the full path/filename of your input file.
book = load_workbook('Test Input.xlsx')
#Takes user input for the no. of sheets on which this change needs to applied.
print('Enter no. of sheets:')
n = int(input())
for x in range(0,n):
  sheet = book.get_sheet_names()[x]
  ws = book.get_sheet_by_name(sheet)
  i=0
  for row in ws.iter_rows():
    i+=1
    for cell in row:
        if cell.font.strike:
            ws.delete_rows(i)
            ws.insert_rows(i)
#Saving the output as a separate file within the same path as input file.
book.save('Output.xlsx')


"""
For any queries feel free to reach out to me on anmolraj006@gmail.com
or ping me over linkedin: https://www.linkedin.com/in/anmol-raj-/
"""